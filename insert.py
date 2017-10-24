#!/usr/bin/python
# -*- coding: utf-8 -*-

import pickle
import openpyxl
import re
import ipaddr
import MySQLdb
import rtapi
from functools import partial

def AssignAnywhereChassisSlot(self,chassis_name,slot_number,server_name, object_tid):
    '''Assign server objects to server chassis'''
    chassis_id = self.GetObjectId(chassis_name)
    server_id = self.GetObjectId(server_name)
    slot_attribute_id = self.GetAttributeId("Slot number")

    sql = "SELECT string_value FROM AttributeValue WHERE object_id = %d AND object_tid = %d AND attr_id = %d" % (server_id, object_tid, slot_attribute_id)
    result = self.db_query_one(sql)

    if result != None:
        # Try to update Value, if no success Insert new value
        sql = "UPDATE AttributeValue SET string_value = '%s' WHERE object_id = %d AND object_tid = %d AND attr_id = %d" % (slot_number, server_id, object_tid, slot_attribute_id)
        self.db_insert(sql)
    else:
        # Assign slot number to server
        sql = "INSERT INTO AttributeValue (object_id,object_tid,attr_id,string_value) VALUES ( %d, %d, %d, '%s')" % ( server_id, object_tid, slot_attribute_id, slot_number)
        print server_id
        self.db_insert(sql)


    # Assign server to chassis
    # Check if it's connected
    sql = "SELECT parent_entity_id FROM EntityLink WHERE child_entity_type = 'object' AND child_entity_id = %d" % (server_id)
    result = self.db_query_one(sql)

    if result != None:
    # Object is connected to someone
        if result[0] != chassis_id:
        # Connected to differend chassis/chassis
            sql = "UPDATE EntityLink SET parent_entity_id = %d WHERE child_entity_id = %d AND child_entity_type = 'object' AND parent_entity_id = %d" % (chassis_id, server_id, result[0])
            self.db_insert(sql)

            old_object_name = self.GetObjectName(result[0])
            self.InsertLog(old_object_name, "Unlinked server %s" % (server_name))
            self.InsertLog(server_id, "Unlinked from Blade Chassis %s" % (old_object_name))
            self.InsertLog(chassis_id, "Linked with server %s" % (server_name))
            self.InsertLog(server_id, "Linked with Blade Chassis %s" % (chassis_name))

    else:
    # Object is not connected
        sql = "INSERT INTO EntityLink (parent_entity_type, parent_entity_id, child_entity_type, child_entity_id) VALUES ('object', %d, 'object', %d)" % (chassis_id, server_id)
        self.db_insert(sql)
        self.InsertLog(chassis_id, "Linked with server %s" % (server_name))
        self.InsertLog(server_id, "Linked with Blade Chassis %s" % (chassis_name))

def AssignServerUnit(self,rack_name,unit_number,server_name):
    '''Assign server objects to server chassis'''
    unit_number = re.sub(r'\s', '', str(unit_number))
    #numbers = unit_number.split(",")
    numbers = re.split(r'[.,]',unit_number)
    atoms = ['front', 'interior', 'rear']
    reorder = []
    for el in numbers:
	#print el
	m = re.match('^([\d]*)-([\d]*)$', el)
        if m:
    	    #numbers.remove(el)
	    for x in range(int(m.group(1)), int(m.group(2))+1):
    		reorder.append(str(x))
    	else:
    	    reorder.append(el)

    rack_id = self.GetObjectId(rack_name)
    server_id = self.GetObjectId(server_name)
    #print rack_id
    #print server_id
    #print numbers
    for unit in reorder:
	#sql = "SELECT object_id FROM RackSpace where rack_id= %d AND unit_no = '%s'" % (rack_id, unit)
	sql = "SELECT O.name, O.objtype_id FROM RackSpace AS RS LEFT JOIN Object AS O ON RS.object_id=O.id where RS.rack_id= %d AND unit_no = '%s' limit 1" % (rack_id, unit)
	result = self.db_query_one(sql)
	if result == None:
	    for atom in atoms:
        	sql = "INSERT INTO RackSpace (rack_id, unit_no, atom, state, object_id) VALUES ('%s', %d, '%s', 'T', %d)" % (rack_id, int(unit), atom, server_id)
        	self.db_insert(sql)
        else:
    	    return result
        
    return None
    
def UpdateNetworkInterfaceMAC(self,object_id,interface,mac):
    '''Add network interfece to object if not exist'''
    sql = "SELECT id,name FROM Port WHERE object_id = %d AND name = '%s'" % (object_id, interface)
    result = self.db_query_one(sql)
    if result == None:
        sql = "INSERT INTO Port (object_id,name,iif_id,type,l2address) VALUES (%d,'%s',1,24,'%s')" % (object_id,interface,mac)
        self.db_insert(sql)
        port_id = self.db_fetch_lastid()
    else:
	port_id = self.db_fetch_lastid()
    return port_id

def MatchAttributeId(self,searchstring):
    '''Search racktables database and get attribud id based on search string as argument'''
    sql = "SELECT id FROM Attribute WHERE name='"+searchstring+"'"
    result = self.db_query_one(sql)
    if result != None:
       getted_id = result[0]
    else:
	getted_id = self.GetAttributeId(searchstring)
    return getted_id

def AddRackObject(rt, ecell, scell):
    #Insert New Rack Object
    #hostname = ecell[61]
    hostname = ecell[scell.index(("Имя хоста").decode('utf8'))]
    service_purpose = ecell[scell.index(("Назначение").decode('utf8'))]

    if (hostname == 'N/A' or hostname == 'N/A' or hostname == 'N/A' or hostname == ('не применимо').decode('utf8') or hostname == ('не установлено').decode('utf8')):
	hostname = ecell[scell.index(("№ КЭ").decode('utf8'))]
    
    #Action during process
    action = 0
    
    if rt.ObjectExistName(hostname):
	object_td = rt.GetObjectId(hostname)
	file_ke = ecell[scell.index(("№ КЭ").decode('utf8'))]
	if rt.GetAttributeValue(object_td, rt.GetAttributeId("№ КЭ")):
	    stored_ke = rt.GetAttributeValue(object_td, rt.GetAttributeId("№ КЭ"))[0]
	    if file_ke == stored_ke:
		#Already added
		action = 0
		return "Rack Object " + ecell[scell.index(("Имя хоста").decode('utf8'))] + " already exist in " + stored_ke
	    else:
		#It means hostname duplicate
		hostname = hostname+"-1"
		if rt.ObjectExistName(hostname):
		    action = 0
		else:
		    action = 1
	else:
	    #Probably error in inserting - return operation
	    action = 0
    else:
	#Create new object
	action = 1
	
    print hostname
    if rt.GetDictionaryId(ecell[2]):
        server_type_id = rt.GetDictionaryId(ecell[2])
    else:
        sql = "INSERT INTO Dictionary (chapter_id, dict_sticky, dict_value) VALUES (%d, 'yes', '%s')" % (1, ecell[2])
        rt.db_insert(sql)
        print ecell[2] + " added to dictionary"
        server_type_id = rt.GetDictionaryId(ecell[2])
        raw_input("Press Enter to continue...")

    #Adding Rack Object
    if action == 1:
        rt.AddObject(hostname, server_type_id, hostname, hostname)

    added = 0
    object_id = rt.GetObjectId(hostname)

    #Adding General Attribute
    for Attr in scell:
        cattr = re.sub(r'\s+', ' ', Attr.encode('utf8'))
        #Adding Device model
        if cattr == "Модель":
    	    sql = "SELECT dict_key FROM Dictionary WHERE dict_value = '%s'" % (ecell[scell.index(Attr)])
    	    result = rt.db_query_one(sql)
    	    print result
	    #model_id = rt.GetDictionaryId(ecell[scell.index(Attr)])
	    if result == None:
		#    if server_type_id == 4:
		#	model_type_id = 11
		#    elif server_type_id == 1502:
		#	model_type_id = 31
		#    elif server_type_id == 1504:
		#	model_type_id = 33
		#    elif server_type_id == 50091:
		#	model_type_id = 24
		#    elif server_type_id == 50124:
		#	model_type_id = 18
		#    elif server_type_id == 50126:
		#	model_type_id = 18
		#    else:
		#	model_type_id = 12
		
	        sql = "INSERT INTO Dictionary (chapter_id, dict_sticky, dict_value) SELECT chapter_id, 'yes', '%s' FROM racktables_db.AttributeMap where attr_id=2 and objtype_id=%d" % (ecell[scell.index(Attr)], server_type_id)
	        #sql = "INSERT INTO Dictionary (chapter_id, dict_sticky, dict_value) VALUES (%d, 'yes', '%s')" % (model_type_id, ecell[scell.index(Attr)])
	        raw_input("Press Enter to continue...")
	        rt.db_insert(sql)
	    model_id = rt.GetDictionaryId(ecell[scell.index(Attr)])
	    print "Модель " + str((ecell[scell.index(Attr)]).encode('utf8')) + " "+ str(model_id)
	    attr_id  = int(rt.MatchAttributeId(Attr));
	    sql = "SELECT uint_value FROM AttributeValue WHERE object_id = %d AND attr_id = %d AND object_tid = %d" % (object_id, attr_id, server_type_id)
	    result = rt.db_query_one(sql)
	    if result == None:
	        sql = "INSERT INTO AttributeValue (uint_value, object_id, attr_id,object_tid) VALUES (%d, %d, %d, %d)" % (model_id, object_id, attr_id, server_type_id)
	    else:
	        sql = "UPDATE AttributeValue SET uint_value = %d WHERE object_id = %d AND attr_id = %d AND object_tid = %d" % (model_id, object_id, attr_id, server_type_id)
	    print sql
	    rt.db_insert(sql)
	    raw_input("Press Enter to continue...")
	#Plecement Server to Rack
	elif cattr == "Юнит №":
		#racks = ecell[scell.index(("Стойка").decode('utf8'))].split("-")
		racks = re.split(r'[-.,]',ecell[scell.index(("Стойка").decode('utf8'))])
		for rack in racks:
		    rack = rack.encode('utf8')
		    if re.match('\d', str(ecell[scell.index(Attr)])):
			print "Размещено"
			if rt.GetObjectId(rack):
			    print "Стойка " + str(rack)
			    print "unit " + str(ecell[scell.index(Attr)])
			else:
			    m = re.match('[A-Z]', rack)
			    if m:
				if rt.GetObjectId(m.group(0)):
				    sql = "INSERT INTO Object (name,objtype_id) VALUES ('%s',%d)" % (rack,1560)
				    rt.db_insert(sql)
				    rack_id = rt.GetObjectId(rack)
				    sql = "INSERT INTO EntityLink (parent_entity_type,parent_entity_id,child_entity_type,child_entity_id) VALUES ('row',%d,'rack',%d)" % (int(rt.GetObjectId(m.group(0))),int(rack_id))
				    rt.db_insert(sql)
				    sql = "INSERT INTO AttributeValue (object_id,object_tid,attr_id,uint_value) VALUES (%d,%d,%d,%d)" % (int(rack_id),1560,27,41)
				    rt.db_insert(sql)
				    sql = "INSERT INTO AttributeValue (object_id,object_tid,attr_id,uint_value) VALUES (%d,%d,%d,%d)" % (int(rack_id),1560,29,1)
				    rt.db_insert(sql)
				    print "Создана стойка " + str(rack_id)
			    else:
				print "Такой стойки нет " + str(rack)
				continue
			assign = rt.AssignServerUnit(rack,ecell[scell.index(Attr)],hostname)
			if assign != None:
			    if assign[1] == 1502 and (server_type_id == 4 or server_type_id == 50093 or server_type_id == 50137):
				print "Наш cервер сидит в шасси " + str(assign[0])
				sql = "SELECT child_entity_id FROM EntityLink WHERE child_entity_type = 'object' AND parent_entity_id=%d" % (rt.GetObjectId(assign[0]))
				result = rt.db_query_all(sql)
				chassis_count = 1
				if result != None:
				    chassis_count = len(result)+1
				    slot_number = rt.GetAttributeValue(object_id, rt.GetAttributeId("Slot number"))
				    if slot_number != None:
					rt.AssignChassisSlot(assign[0], slot_number[0], hostname)
					print "Уже добавлено в шасси слот: " + str(slot_number[0])
				    else:
					rt.AssignAnywhereChassisSlot(assign[0], chassis_count, hostname, server_type_id)
					print "Добавлено в шасси: " + str(assign[0])
				else:
				    print "Добавляем в шасси слот: " + str(chassis_count)
				    rt.AssignChassisSlot(assign[0], chassis_count, hostname)

			    elif assign[1] != None and assign[0] != hostname:
				print "Ошибка добавления сервера в шасси " + str(assign)
			    else:
				print "Уже добавлено " + str(assign[0])
		    else:
			ke_num = (rt.GetAttributeValue(object_id, rt.GetAttributeId("№ КЭ")))[0]
			#sql = "SELECT object_id FROM racktables_db.AttributeValue Where string_value='%s'" % (ke_num.split('-')[0])
			sql = "SELECT object_id FROM racktables_db.AttributeValue Where string_value LIKE '%s%%' and string_value not LIKE '%%-%%'" % (ke_num.split('-')[0])
			result = rt.db_query_one(sql)
			#parent_id = result[0]
			if result != None and server_type_id == 1504:
			    rt.LinkVirtualHypervisor(result[0], object_id)
			    print "Мы виртуальная машина"
			else:
			    print "Не размещено: Ошибочные данные размещения"
		#raw_input("Press Enter to continue...")
		rt.InsertAttribute(object_id,server_type_id,10083,ecell[scell.index(("Комната").decode('utf8'))],"NULL",hostname)
	#Adding Ports count
	elif cattr == "Всего" and added == 0:
		print "Кабели"
		rt.InsertAttribute(object_id,server_type_id,10079,ecell[scell.index(Attr)],"NULL",hostname)
		rt.InsertAttribute(object_id,server_type_id,10080,ecell[scell.index(Attr)+1],"NULL",hostname)
		rt.InsertAttribute(object_id,server_type_id,10081,ecell[scell.index(Attr)+2],"NULL",hostname)
		rt.InsertAttribute(object_id,server_type_id,10082,ecell[scell.index(Attr)+3],"NULL",hostname)
		added = 1
	#Adding IP information
	elif cattr == "Имя хоста":
		#Если действительно есть адрес
		hostname_index = int(scell.index(Attr))
		if re.match('\d*\.', ecell[hostname_index+1]):
		    print "HOSTNAME " + ecell[hostname_index] 
		    #ip_addr = (ecell[hostname_index+1]).split("-")[0]
		    ip_addr = (re.split(r'[-,]',(ecell[hostname_index+1]).decode('utf8')))[0]
		    rt.UpdateNetworkInterfaceMAC(object_id,ecell[hostname_index],ecell[hostname_index+4])
		    rt.InterfaceAddIpv4IP(object_id, ecell[hostname_index], ip_addr)
		    rt.InsertAttribute(object_id,server_type_id,10000,ecell[hostname_index+1],"NULL",hostname)
		
		scell[hostname_index] = str("Имяхоста").decode('utf8')
		scell[hostname_index+1] = str("IPaddress").decode('utf8')
	#Exception Адрес
	elif cattr == "Адрес":
		print 
	#adding Other Attribute
	elif rt.MatchAttributeId(Attr):
		attr_id  = int(rt.MatchAttributeId(Attr));
		#print " Тип сервера: " + str(server_type_id)
		print "Attribute " + Attr + " '" + str(attr_id) + "'"
		#+ " Значение: " + ecell[scell.index(Attr)] 
		rt.InsertAttribute(object_id,server_type_id,attr_id,ecell[scell.index(Attr)],"NULL",hostname)
	#Not Used Attribute
	else:
		print "No Attribute: " + Attr
	#attr_id = rt.GetAttributeId("№ КЭ")
        #attr_id = rt.GetAttributeId("Прошивка")
	#print ("Attr ID "+str(attr_id))
        #rtobject.InsertAttribute(object_id,4,attr_id,"NULL",os_id,hostname)
	#print rt.GetObjectId("E3")
        #print rt.GetObjectId("Тестовый сервер")
	#rt.AssignServerUnit("E3",'1-9,16,20-22,25',"Тестовый сервер")
    return "Added " + hostname

# Create connection to database
try:
    # Create connection to database
    db = MySQLdb.connect(host='127.0.0.1',port=3306, passwd='RackPassw0rd',db='racktables_db',user='racktables_user',charset='utf8', init_command='SET NAMES UTF8')
except MySQLdb.Error ,e:
    print "Error %d: %s" % (e.args[0],e.args[1])
    sys.exit(1)

with open ('outfile', 'rb') as fp:
    ecell = pickle.load(fp)
with open ('sfile', 'rb') as fp:
    scell = pickle.load(fp)

# Initialize rtapi with database connection
print("Initializing RT Api object")
rt = rtapi.RTObject(db)
rt.AssignServerUnit = partial(AssignServerUnit, rt);
rt.MatchAttributeId = partial(MatchAttributeId, rt);
rt.UpdateNetworkInterfaceMAC = partial(UpdateNetworkInterfaceMAC, rt)
rt.AssignAnywhereChassisSlot = partial(AssignAnywhereChassisSlot, rt)

#Main part ADD RACK OBJECT
#print AddRackObject(rt,ecell, scell)

# List all objects from database
print rt.ListObjects()
#exit()

# Parsing from Excel
wb = openpyxl.load_workbook(filename = 'gisgmp.xlsx', read_only=True)
print "Excel loaded successfully"
for sheetName in wb.get_sheet_names():
#    if re.match(("Сервера|СЕРВЕРА").decode('utf8'), sheetName, re.IGNORECASE):
    if re.match(("Сети|СЕТИ").decode('utf8'), sheetName, re.IGNORECASE):
#    if re.match(("Соби|СОБИ").decode('utf8'), sheetName, re.IGNORECASE):
#    if re.match(("СХД и СРК").decode('utf8'), sheetName, re.IGNORECASE):
	print "Work with shet " + sheetName
	sheet = wb[sheetName]
        scell = []
	for row in sheet.iter_rows(min_row=4, max_row=4):
    	    for cell in row:
    		if cell.value is not None:
		    scell.append(cell.value)
		else:
		    if sheet.cell(row=2, column=cell.column).value is not None:
    			scell.append(sheet.cell(row=2, column=cell.column).value)
    		    else:
    			scell.append(sheet.cell(row=1, column=cell.column).value)
	    with open('sfile', 'wb') as fp:
		pickle.dump(scell, fp)
	i = 5
	for row in sheet.iter_rows(min_row=5):
	    print row[0].value
	    if row[0].value == None:
		break
    	    ecell = []
	    for cell in row:
		ecell.append(cell.value)
	    with open('outfile', 'wb') as fp:
		pickle.dump(ecell, fp)
		
	    with open ('outfile', 'rb') as fp:
	        ecell = pickle.load(fp)
	    with open ('sfile', 'rb') as fp:
	        scell = pickle.load(fp)

	    print AddRackObject(rt,ecell, scell)

	    raw_input("Row %d Inserted..." % i)
	    i = i + 1

