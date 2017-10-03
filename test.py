#!/usr/bin/python
# -*- coding: utf-8 -*-

import pickle
import openpyxl
import re
import ipaddr
import MySQLdb
import rtapi
from functools import partial

def AssignServerUnit(self,rack_name,unit_number,server_name):
    '''Assign server objects to server chassis'''
    numbers = unit_number.split(",")
    atoms = ['front', 'interior', 'rear']
    for el in numbers:
	m = re.match('^([\d]*)-([\d]*)$', el)
        if m:
    	    numbers.remove(el)
	    for x in range(int(m.group(1)), int(m.group(2))+1):
    		numbers.append(str(x))

    rack_id = self.GetObjectId(rack_name)
    server_id = self.GetObjectId(server_name)
    print rack_id
    print server_id
    print numbers
    for unit in numbers:
	for atom in atoms:
            sql = "INSERT INTO RackSpace (rack_id, unit_no, atom, state, object_id) VALUES ('%s', %d, '%s', 'T', %d)" % (rack_id, int(unit), atom, server_id)
            self.db_insert(sql)
            
'''
def get_server_type_id(self, server_type):
    # Assign server type id to various hardware
    if (ecell[1] == "Сервер х86"): server_id = 4
    elif (ecell[1] == "Серверное шасси"): server_id = 1502
    elif (ecell[1] == "Виртуальная машина"): server_id = 1504
    elif (ecell[1] == "IBM System P"): server_id = 
    elif (ecell[1] == "Консоль (KVM)"): server_id =
    elif (ecell[1] == "Межсетевой экран"): server_id =
    elif (ecell[1] == "Коммутатор управляемый (L3)"): server_id =
    elif (ecell[1] == "Модем"): server_id = 13
    elif (ecell[1] == "Система мониторинга"): server_id =
    elif (ecell[1] == "Маршрутизатор (router)"): server_id =
    elif (ecell[1] == "Оптимизатор потока данных (WAAS)"): server_id =
    elif (ecell[1] == "Крипто-шлюз"): server_id =
    
    return server_id
'''

def AddRackObject(rt, ecell, scell):
    #Insert New Rack Object
    hostname = ecell[61]
    if rt.ObjectExistName(hostname):
	return "Rack Object " + ecell[61]+ " already exist"
    else:
	if rt.GetDictionaryId(ecell[2]):
	    server_type_id = rt.GetDictionaryId(ecell[2])
	else:
	    sql = "INSERT INTO Dictionary (chapter_id, dict_sticky, dict_value) VALUES (%d, 'yes', '%s')" % (1, ecell[2])
	    rt.db_insert(sql)
	    print ecell[2] + " added to dictionary"
	    server_type_id = rt.GetDictionaryId(ecell[2])
	#Adding Rack Object
	rt.AddObject(ecell[61], server_type_id, ecell[3], ecell[61])
	object_id = rt.GetObjectId(hostname)
	for Attr in scell:
	    if rt.GetAttributeId(Attr):
		attr_id  = int(rt.GetAttributeId(Attr));
		print "Attribute " + Attr + " " + str(rt.GetAttributeId(Attr))
		print ecell[scell.index(Attr)]
		rt.InsertAttribute(object_id,server_type_id,attr_id,ecell[scell.index(Attr)],"NULL",hostname)
		exit()
	    else:
		print "No Attribute: " + Attr
	    attr_id = rt.GetAttributeId("№ КЭ")
        attr_id = rt.GetAttributeId("Прошивка")
	print ("Attr ID "+str(attr_id))
        #rtobject.InsertAttribute(object_id,4,attr_id,"NULL",os_id,hostname)
	#print rt.GetObjectId("E3")
        #print rt.GetObjectId("Тестовый сервер")
	#rt.AssignServerUnit("E3",'1-9,16,20-22,25',"Тестовый сервер")
        return "Added " + ecell[61]

# Create connection to database
try:
    # Create connection to database
    db = MySQLdb.connect(host='127.0.0.1',port=3306, passwd='RackPassw0rd',db='racktables_db',user='racktables_user',charset='utf8', init_command='SET NAMES UTF8')
except MySQLdb.Error ,e:
    print "Error %d: %s" % (e.args[0],e.args[1])
    sys.exit(1)

with open ('outfile', 'rb') as fp:
    ecell = pickle.load(fp)
with open ('sfile', 'rb') as fp:
    scell = pickle.load(fp)
print scell[0]

# Initialize rtapi with database connection
print("Initializing RT Api object")
rt = rtapi.RTObject(db)
rt.AssignServerUnit = partial(AssignServerUnit, rt);

print AddRackObject(rt,ecell, scell)

# List all objects from database
print rt.ListObjects()
exit()

# Parsing from Excel
wb = openpyxl.load_workbook(filename = 'rvc.xlsx', read_only=True)
print "Excel loaded successfully"
for sheetName in wb.get_sheet_names():
    if re.match(("Сервера|СЕРВЕРА").decode('utf8'), sheetName, re.IGNORECASE):
	print "Work with sheet " + sheetName
	sheet = wb[sheetName]
        scell = []
	for row in sheet.iter_rows(min_row=4, max_row=4):
    	    for cell in row:
    		if cell.value is not None:
		    scell.append(cell.value)
		else:
		    if sheet.cell(row=2, column=cell.column).value is not None:
    			scell.append(sheet.cell(row=2, column=cell.column).value)
    		    else:
    			scell.append(sheet.cell(row=1, column=cell.column).value)
	    with open('sfile', 'wb') as fp:
		pickle.dump(scell, fp)
	for row in sheet.iter_rows(min_row=5, max_row=5):
    	    ecell = []
	    for cell in row:
		ecell.append(cell.value)
	    with open('outfile', 'wb') as fp:
		pickle.dump(ecell, fp)

